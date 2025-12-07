import re
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict
from datetime import datetime

class ExcelProcessor:
    """Класс для чтения и обработки исходного Excel-файла."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.df: Optional[pd.DataFrame] = None

    def process_file(self) -> Tuple[List[Dict[str, Any]], str]:
        """
        Основной метод, который загружает и парсит файл.
        Возвращает кортеж из списка товаров и информации об отгрузке.
        """
        self._load_dataframe()
        shipment_info = self._extract_shipment_details()
        orders = self._parse_orders()
        
        # Преобразуем информацию об отгрузке в строку для совместимости с GUI
        shipment_info_str = f"Отгрузка №{shipment_info['number']} от {shipment_info['date']}"
        
        return orders, shipment_info_str

    def _load_dataframe(self):
        """Загружает данные из Excel-файла в DataFrame."""
        if self.df is None:
            try:
                self.df = pd.read_excel(
                    self.file_path,
                    header=4, # Строка 5 (индекс 4)
                    dtype=str,
                    engine='openpyxl'
                )
            except Exception as e:
                raise ValueError(f"Не удалось прочитать файл Excel: {e}")

    def _extract_shipment_details(self) -> dict:
        """Извлекает номер и дату отгрузки из первой строки."""
        # Читаем первую строку отдельно, так как основной df теперь с header=3
        try:
            temp_df = pd.read_excel(
                self.file_path,
                header=None,
                nrows=1,
                dtype=str,
                engine='openpyxl'
            )
        except Exception:
             return {
                "number": f"SHIP_{datetime.now().strftime('%Y%m%d_%H%M')}",
                "date": datetime.now().strftime('%d-%m-%Y')
            }

        if temp_df.empty:
             return {
                "number": f"SHIP_{datetime.now().strftime('%Y%m%d_%H%M')}",
                "date": datetime.now().strftime('%d-%m-%Y')
            }

        first_row_text = " ".join(str(cell) for cell in temp_df.iloc[0, :].dropna())
        pattern = re.compile(r'№\s*([\w.-]+)\s+от\s+([\d.]+)')
        match = pattern.search(first_row_text)
        
        if match:
            shipment_number = match.group(1)
            date_str = match.group(2)
            try:
                date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                formatted_date = date_obj.strftime('%d-%m-%Y')
            except ValueError:
                formatted_date = datetime.now().strftime('%d-%m-%Y')
            return {"number": shipment_number, "date": formatted_date}

        return {
            "number": f"SHIP_{datetime.now().strftime('%Y%m%d_%H%M')}",
            "date": datetime.now().strftime('%d-%m-%Y')
        }

    def _parse_orders(self) -> List[Dict[str, Any]]:
        """Преобразует DataFrame в список товаров для сборки."""
        if self.df is None:
            raise ValueError("DataFrame не загружен.")
        
        # Очищаем названия колонок от пробелов
        self.df.columns = self.df.columns.astype(str).str.strip()

        required = {"Наименование товара", "Количество", "Артикул"}
        # "Ячейка" и "Штрихкод" могут называться по-разному или отсутствовать в явном виде,
        # но мы ожидаем их наличие согласно ТЗ. Добавим проверку.
        
        if not required.issubset(set(self.df.columns)):
            missing = required - set(self.df.columns)
            raise ValueError(f"Отсутствуют обязательные колонки: {', '.join(missing)}")

        orders = []
        for _, row in self.df.iterrows():
            try:
                # Пропускаем пустые строки или строки без наименования
                if pd.isna(row["Наименование товара"]) or not str(row["Наименование товара"]).strip():
                    continue

                quantity = int(float(row["Количество"]))
                if quantity <= 0:
                    continue
                
                # Извлекаем новые поля
                location = str(row.get("Ячейка", "")).strip()
                if location == "nan": location = ""
                
                barcode = str(row.get("Штрихкод", "")).strip()
                if barcode == "nan": barcode = ""

                orders.append({
                    "name": str(row["Наименование товара"]).strip(),
                    "quantity": quantity,
                    "article": str(row["Артикул"]).strip() or "?",
                    "location": location,
                    "barcode": barcode
                })
            except (ValueError, KeyError, TypeError):
                # Пропускаем строки с некорректными данными
                continue
        
        if not orders:
            raise ValueError("Не найдено валидных позиций для сборки.")

        return orders


class ExcelWriter:
    """Класс для генерации итогового Excel-файла."""

    def __init__(self, collected_data: List[Dict], shipment_info: str, discrepancies: List[str], original_file_path: str):
        self.collected_data = collected_data
        self.shipment_info = shipment_info
        self.discrepancies = discrepancies
        self.original_file_path = Path(original_file_path)

    def generate_final_file(self) -> str:
        """
        Создает итоговый Excel-файл и возвращает его имя.
        """
        # Группируем данные по коробкам
        boxes = defaultdict(list)
        for record in self.collected_data:
            boxes[record['box']].append(record)

        if not boxes:
            raise ValueError("Нет данных для записи.")

        # Создаем новый Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результат Сборки"

        # Добавляем информацию об отгрузке и расхождениях
        ws.cell(row=1, column=1, value=self.shipment_info)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        
        if self.discrepancies:
            start_row = 3
            ws.cell(row=start_row, column=1, value="Расхождения:")
            for i, note in enumerate(self.discrepancies, 1):
                ws.cell(row=start_row + i, column=1, value=note)

        # Определяем начальную строку для данных о коробках
        data_start_row = 3 + len(self.discrepancies) + 2 if self.discrepancies else 4
        current_col = 1

        # Записываем данные по каждой коробке
        for box_num in sorted(boxes.keys()):
            # Заголовок коробки
            header_cell = ws.cell(row=data_start_row, column=current_col, value=f"КОРОБКА №{box_num}")
            ws.merge_cells(start_row=data_start_row, start_column=current_col, end_row=data_start_row, end_column=current_col + 2)
            header_cell.alignment = Alignment(horizontal="center")
            
            # Заголовки столбцов (Обновленный порядок)
            ws.cell(row=data_start_row + 1, column=current_col, value="Кол-во")
            ws.cell(row=data_start_row + 1, column=current_col + 1, value="Артикул")
            ws.cell(row=data_start_row + 1, column=current_col + 2, value="Штрихкод") # Вместо Названия

            # Данные
            row_cursor = data_start_row + 2
            for record in boxes[box_num]:
                ws.cell(row=row_cursor, column=current_col, value=record['quantity'])
                ws.cell(row=row_cursor, column=current_col + 1, value=record['article'])
                # Используем штрихкод, если он есть, иначе пусто или имя (если штрихкода нет, но просили штрихкод)
                # В ТЗ сказано "Штрихкод (без Наименования товара)"
                ws.cell(row=row_cursor, column=current_col + 2, value=record.get('barcode', '')) 
                row_cursor += 1
            
            # Сдвигаем курсор для следующей коробки
            current_col += 4 

        # Генерируем имя выходного файла
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        output_filename = self.original_file_path.stem + f"_сборка_{timestamp}.xlsx"
        output_path = self.original_file_path.parent / output_filename
        
        wb.save(output_path)
        return str(output_path)